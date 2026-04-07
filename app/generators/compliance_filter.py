"""Compliance Filter Generator (Excel output).

Generates an engineering specification compliance table matching the sample format:
- Sheet "附件8-1 電,水,藥,排氣": Categorized spec table with merged category cells
- Categories: 冰水, 電力, 潔淨室, 規劃, 工程施工要求
- Each category has sub-types (基本要求, 重點要求, 原則, 一般性共同要求, 權責/安全管制說明)
- Columns: 項目 | 類別 | 規範說明 | 本項次不適用 | 驗收結果 | 驗收日期 | 驗收者
"""

import io
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.generators.base import BaseGenerator
from app.llm.batch_processor import LLMBatchProcessor
from app.models.quote_data import ComplianceResult, QuoteData

logger = logging.getLogger(__name__)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
CATEGORY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

SPEC_CONTENT_START_ROW = 2


class ComplianceFilterGenerator(BaseGenerator):
    doc_type = "compliance_filter"
    output_extension = "xlsx"

    async def generate(self, quote_data: QuoteData, **kwargs) -> bytes:
        """Generate filtered compliance specification Excel file."""
        compliance_results = kwargs.get("compliance_results")

        if self.template_path and self.template_path.exists():
            return await self._filter_template(quote_data, compliance_results)
        else:
            return await self._generate_from_scratch(quote_data, compliance_results)

    async def _filter_template(
        self, quote_data: QuoteData,
        compliance_results: list[ComplianceResult] | None,
    ) -> bytes:
        """Filter an existing compliance spec template."""
        from app.templates.engine import ExcelTemplateEngine

        try:
            engine = ExcelTemplateEngine(self.template_path)
        except Exception as e:
            logger.warning("Cannot open template %s: %s, generating from scratch", self.template_path, e)
            return await self._generate_from_scratch(quote_data, compliance_results)
        ws_name = engine.get_sheet_names()[0]

        all_rows = engine.get_all_rows(ws_name, start_row=SPEC_CONTENT_START_ROW)

        if not compliance_results and self.llm:
            spec_items = []
            for i, row in enumerate(all_rows):
                content = " | ".join(str(cell) for cell in row if cell is not None)
                if content.strip():
                    spec_items.append({"row_index": i, "content": content[:200]})

            compliance_results = await self.llm.generate_compliance_judgments(
                quote_data, spec_items
            )

        rows_to_delete = []
        if compliance_results:
            irrelevant_indices = {
                r.row_index for r in compliance_results if not r.is_relevant
            }
            rows_to_delete = [
                idx + SPEC_CONTENT_START_ROW for idx in irrelevant_indices
            ]

        engine.delete_rows(ws_name, rows_to_delete)
        return engine.save_to_bytes()

    async def _generate_from_scratch(
        self, quote_data: QuoteData,
        compliance_results: list[ComplianceResult] | None,
    ) -> bytes:
        """Generate a complete engineering spec table from scratch."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "附件8-1 電,水,藥,排氣"

        meta = quote_data.metadata

        # Page setup
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1

        # Column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 65
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 10

        # Row 1: Title
        ws.merge_cells("A1:G1")
        ws["A1"] = "工  程  規  範  表(一)"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Row 2: Headers
        headers = ["項目", "類別", "規          範          說          明",
                    "本項次不適用", "驗收結果", "驗收日期", "驗收者"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Generate spec items based on quote content
        spec_categories = _infer_spec_categories(quote_data)

        # Use LLM results to filter if available
        if compliance_results:
            relevant_indices = {r.row_index for r in compliance_results if r.is_relevant}
        else:
            relevant_indices = None  # Keep all

        current_row = 3
        for cat_name, cat_type, specs in spec_categories:
            if not specs:
                continue

            # Filter specs if LLM results available
            if relevant_indices is not None:
                filtered = [(idx, s) for idx, s in enumerate(specs) if idx in relevant_indices]
                if not filtered:
                    continue
                specs_to_write = [s for _, s in filtered]
            else:
                specs_to_write = specs

            cat_start_row = current_row
            type_start_row = current_row

            for spec_idx, spec_text in enumerate(specs_to_write):
                # Column C: Spec text
                ws.cell(row=current_row, column=3).value = f"{spec_idx + 1}. {spec_text}"
                ws.cell(row=current_row, column=3).font = Font(size=10)
                ws.cell(row=current_row, column=3).alignment = Alignment(wrap_text=True, vertical="top")

                # Columns D-G: empty checkoff columns
                for col in range(4, 8):
                    ws.cell(row=current_row, column=col).border = THIN_BORDER

                ws.cell(row=current_row, column=1).border = THIN_BORDER
                ws.cell(row=current_row, column=2).border = THIN_BORDER
                ws.cell(row=current_row, column=3).border = THIN_BORDER

                current_row += 1

            cat_end_row = current_row - 1

            # Merge category cells (Column A)
            if cat_end_row > cat_start_row:
                ws.merge_cells(f"A{cat_start_row}:A{cat_end_row}")
            ws.cell(row=cat_start_row, column=1).value = cat_name
            ws.cell(row=cat_start_row, column=1).font = Font(bold=True, size=10)
            ws.cell(row=cat_start_row, column=1).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            ws.cell(row=cat_start_row, column=1).fill = CATEGORY_FILL

            # Merge type cells (Column B)
            if cat_end_row > type_start_row:
                ws.merge_cells(f"B{type_start_row}:B{cat_end_row}")
            ws.cell(row=type_start_row, column=2).value = cat_type
            ws.cell(row=type_start_row, column=2).font = Font(size=10)
            ws.cell(row=type_start_row, column=2).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Footer notes
        footer_notes = [
            "1. 施工期間安全遵守本公司：1.承攬商管理辦法 2.特殊作業申請 3.現場安全告知，違規者依公司規定予以罰款或扣抵工程款。",
            "2. 本表內容無明確規範的部分則以設備二次配施工規範為準",
            "3. 工程規劃應以合乎設備二次配施工規範的前提下進行，並事先與權責單位充分溝通並出具相關圖面、標單等資料",
            "4. *號標示的項目，適用於台灣地區(大陸地區暫不適用)",
        ]
        for note in footer_notes:
            ws.merge_cells(f"A{current_row}:G{current_row}")
            ws.cell(row=current_row, column=1).value = note
            ws.cell(row=current_row, column=1).font = Font(size=9)
            ws.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True)
            current_row += 1

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def _infer_spec_categories(quote_data: QuoteData) -> list[tuple[str, str, list[str]]]:
    """Infer applicable specification categories from quote content.

    Returns list of (category_name, sub_type, [spec_texts]).
    """
    items = quote_data.get_real_items()
    all_desc = " ".join(item.description for item in items).lower()

    categories = []

    # 冰水 (chilled water)
    if any(kw in all_desc for kw in ["冰水", "冷卻", "配管", "管路", "鋼管", "閥", "保溫", "軟管"]):
        categories.append(("冰水", "基本要求", [
            "一般冰水管路材質為GIP鍍鋅鐵管(2\"以上使用氬焊)並含預留及管末回流閥；附面積流量計、Y型過濾器及流量可調式閥件；保溫使用防火保溫材",
            "二次配管路(管路材質為GIP鍍鋅鐵管外亦得選用PVC SCH80管)上需包含壓力錶、溫度表、流量計等量測儀器",
            "管路下至落點，落點前需加設釋氣閥，回水端附面積流量計；入水端設Y型過濾器",
            "水管路需保溫(含設備端；統一黑色膠帶)，並可於入設備端前加裝bypass迴路，附溫度表、壓力表、流量可調式閥件",
            "冰水全工程橡膠材質一律Viton(含packing及O-ring等)",
            "設備銜接需使用軟管方式時冰水管徑大於1/2\"，需使用高壓管壓接配管，接頭限定使用SUS304等級",
            "容易發生水錘振動的管路於穿孔處(如高架地板)，需以防火保溫材包覆，避免碰觸摩擦造成破裂",
        ]))

    # 電力 (electrical)
    if any(kw in all_desc for kw in ["電", "配電", "線", "ffu", "分電箱", "線架", "配線", "emt", "xlpe"]):
        categories.append(("電力", "基本要求", [
            "電力電纜限用一級廠(太平洋、華新麗華)XLPE(PEX)電纜600V級",
            "ACB&NFB無溶絲開關限用Fuji、士林等一級廠產品",
            "Cable Tray體採鋁製品並烤漆(每一銜接處應該有8 mm2的綠色接地線)",
            "電力電纜不得有接續處；佈於線槽上須以紮線帶固定",
            "Cable Tray於安裝完成後須使用米色epoxy漆，將牙條及脫漆處予以修補",
            "電力電纜(Cable)進入控盤一律由控盤側面或後面下方開孔進入，不得由上方開孔進入",
            "電力電纜不得與設備控制用電纜共處於同一Cable Tray(EMI 防止)",
            "二次配之Cable Tray及電力電纜啟點為一次側電力盤NFB二次側起，至設備端總NFB一次側止",
            "潔淨室內之Cable Tray的固定方式：橫樑部分應以白鐵材質(#304)角鐵固定，其餘以白鐵材質(#304)牙條搭配膨脹螺絲固定",
            "所有電纜必須頭尾加線碼，並標示相序、電源引接處、流水號及色套",
            "電纜架支撐應以4分膨脹螺栓固定，並使用4分牙條；凡經過橫樑處需用角鐵支撐電纜架",
            "進入分電盤及控盤一律由盤體側面或後面下方開孔進入，不得由上方開孔進入",
            "無熔絲開關其規格按規定辦理；銜接電盤為插入方式者，須符合PLUG-IN型無熔絲開關",
            "銅排之電流規格為NFB額定之1.2倍以上",
            "動力設備分路導線其最小線徑不得小於3.5mm，線路不得裸露，應以適當防護",
            "配電結束時，需與業主會同量測線路配置是否正確，再進行送電",
        ]))

    # 潔淨室 (cleanroom)
    if any(kw in all_desc for kw in ["無塵室", "潔淨", "庫板", "t-grid", "高架地板", "ffu"]):
        categories.append(("潔淨室", "重點要求", [
            "潔淨室內施作前確認各項排氣、防護措施是否足夠，施工人員經廠內認證",
            "潔淨室內禁止所有管路或各類管線直接由天花板穿過",
            "潔淨室內的配管需行走於回風牆內，潔淨室作業區須盡量減少管線配置；並禁止直接貼於地面配管",
            "高架地板下所有特殊、一般氣體管線銜接應為無縫焊接及VCR gasket compression seals",
            "高架地板下管架(Support)及所有五金配件皆用SUS304材質",
        ]))

    # 規劃 (planning)
    categories.append(("規劃", "原則", [
        "金屬管配管如採取焊接方式則限用氬焊，焊口一律要酸鹼洗及防鏽處理(潔淨室設置完成後嚴禁動火)",
        "取得生產設備二次配需求資訊；並提供廠內一次供應源SPEC供廠商設計規劃",
        "符合安全的前提下，優先消化廠內閒置物料與堪用的舊品(含管路)以降低成本",
        "二次配管路動線依實際需求規劃整合共同管架，考慮美觀與安全為重點",
        "設備銘牌、管路標示應清楚便於識別，不同流體應以顏色區分，流向以箭頭清楚標示",
        "管路及電線禁止配置在生產機台及機台輔助設施上方",
        "管線（含風管）及電線槽穿牆處需採美觀原則，且必須採用2mm不銹鋼板材內襯防火材",
        "如遇須穿牆處，須以機械鑽孔方式施作，禁止人工打鑿",
        "穿板之管線必須作樓板穿孔處通行，防水氣密填縫及EPOXY加玻纖三層止水防液堤及止水墩",
        "照明燈具應優先採用LED燈具(無法採用時應事先經業主同意)",
        "吊架材質：牙條與角鋼於潔淨室區應採SUS304；走道區得依環境採鍍鋅材質",
    ]))

    # 工程施工要求 - 一般性共同要求
    categories.append(("工程施工要求", "一般性\n共同要求", [
        "五金另件(如膨脹螺絲、螺絲、墊片等)限用SUS304材質；新設之電氣箱體應做好防水措施",
        "完工後須依本公司規範進行管路流向及標示管內物質(管路每3M及轉角兩邊)",
        "各閥門應標示操作狀態(紅色常關、綠色常開、藍色調整)且需有中英文標示",
        "潔淨室內施作時，廠務或設備人員應於施作前確認各項排氣、防護措施是否足夠",
        "工程驗收時確認管線穿牆或庫板時應做好防火填塞",
        "工程驗收時確認管線標示是否清楚，顏色是否符合規定",
        "耐壓試驗為管路系統配管工程完成後所施行之試驗，應以管材最高許可使用壓力之1.5倍水壓試驗並維持1hr以上無洩漏方為合格",
        "保壓試驗應以系統壓力最高許可使用壓力之1.1倍壓力維持24hrs以上無洩漏方為合格",
        "工程施作項目貫穿隔間之空隙需施以膨脹型防火填塞，各類防火材料需有防火材料證明",
        "所有新設帶電線路須以設置線槽為原則",
        "法蘭固定螺栓、螺帽均須附墊圈，支撐牙條與吊架均須附保護套，材質須為耐酸、鹼腐蝕",
        "法蘭各孔洞固定螺栓須全部確實以對角方式上鎖固定，以確保密合不滲漏",
    ]))

    # 工程施工要求 - 權責/安全管制說明
    categories.append(("工程施工要求", "權責/\n安全管制\n說明", [
        "本工程為連工帶料，需附圖面與共同標單供審核參考，實際數量依現場為準",
        "廠務或設備人員於工程規劃時需要求承包商提供完整圖面與標單，並由權責單位完成審核同意",
        "廠務或設備人員應於工程驗收時特別注意是否有使用我方提供之舊品或材料，並依規定做好識別標示",
        "完工需附圖面",
        "施工現場嚴禁煙火、飲食，並遵守業主所有安全作業規定",
        "須配合設備裝機進度完工，施工人員配合加班趕工，不辦追加",
        "工程期間若有修改動線位置或原設計位置點時，百分之五以內不得追加任何款項",
        "成本需含清安費不再追加(共同標單內不應有此項目)",
        "每日施工完後須保持現場清潔",
        "須於施工前完成各項入廠及工安申請，並依安全衛生法相關規定配戴防護具及器械",
        "保固期(完成驗收後)一年內若非人為因素一律無償負責修繕",
        "若因施工不慎導致本廠損失將負責修繕或賠償",
        "若涉及潔淨室案件需遵照潔淨室施工規範或作業辦法施作，施工過程需有適當防護隔離，完工後需進行環境落塵量檢測",
        "依《EA52-005承攬商管理辦法》執行監工並遵守各項承攬商規定",
        "各項管線及閥件標示需符合設備二次配規範及勞工安全設施規則",
        "移動式用電設備應加裝漏電斷路器，避免漏電造成感電傷害",
    ]))

    return categories
