"""Common Bid / Blank Bid Form Generator (Excel output).

Generates a professional bid form matching the sample format:
- Sheet 1 "共同標單": Full hierarchy (壹/一/1.1 levels), section headers,
  formulas =IF(OR(D##="",E##=""),"",D##*E##), subtotals with =SUM(),
  summary rows, vendor signature area
- Sheet 2 "_驗證": Verification sheet with original prices, arithmetic checks
"""

import io
import logging
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.generators.base import BaseGenerator
from app.models.quote_data import QuoteData, QuoteLineItem, QuoteSection

logger = logging.getLogger(__name__)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SECTION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

# Chinese numerals for major sections
CHINESE_NUMERALS = ["壹", "貳", "參", "肆", "伍", "陸", "柒", "捌", "玖", "拾"]
SUB_NUMERALS = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十"]


class CommonBidGenerator(BaseGenerator):
    doc_type = "common_bid"
    output_extension = "xlsx"

    async def generate(self, quote_data: QuoteData, **kwargs) -> bytes:
        """Generate blank bid form Excel file with verification sheet."""
        if self.template_path and self.template_path.exists():
            return self._fill_template(quote_data)
        else:
            return self._generate_from_scratch(quote_data)

    def _fill_template(self, quote_data: QuoteData) -> bytes:
        """Fill an existing Excel template."""
        from app.templates.engine import ExcelTemplateEngine
        engine = ExcelTemplateEngine(self.template_path)
        return engine.save_to_bytes()

    def _generate_from_scratch(self, quote_data: QuoteData) -> bytes:
        """Generate a complete bid form from scratch matching the sample format."""
        wb = openpyxl.Workbook()

        # ======== Sheet 1: 共同標單 ========
        ws = wb.active
        ws.title = "共同標單"
        self._build_bid_sheet(ws, quote_data)

        # ======== Sheet 2: _驗證 ========
        ws2 = wb.create_sheet("_驗證")
        self._build_verification_sheet(ws2, quote_data)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _build_bid_sheet(self, ws, quote_data: QuoteData):
        """Build the main bid sheet."""
        meta = quote_data.metadata

        # Page setup
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_margins.left = 0.6
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.8
        ws.page_margins.bottom = 0.6

        # Column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 12

        # Row 1: Title
        ws.merge_cells("A1:G1")
        c = ws["A1"]
        c.value = "共 同 標 單"
        c.font = Font(size=16, bold=True)
        c.alignment = Alignment(horizontal="center")

        # Row 2-6: Header info
        doc_num = meta.document_number or ""
        project_display = meta.project_name
        if doc_num:
            project_display = f"（{doc_num}）{meta.project_name}"

        header_data = [
            ("A2", "工程名稱", "B2", project_display, "D2", "標單名稱", "E2", "共同標單"),
            ("A3", "報價單號", "B3", doc_num, "D3", "報價日期", "E3",
             str(meta.quote_date).replace("-", "/") if meta.quote_date else ""),
            ("A4", "投標廠商", "B4", "", "D4", "聯絡人", "E4", ""),
            ("A5", "聯絡電話", "B5", "", "D5", "E-mail", "E5", ""),
            ("A6", "地址", "B6", "", "D6", "稅率", "E6", 0.05),
        ]
        for row_data in header_data:
            for i in range(0, len(row_data), 2):
                cell_ref, val = row_data[i], row_data[i + 1]
                ws[cell_ref] = val
                if cell_ref.startswith("A") or cell_ref.startswith("D"):
                    ws[cell_ref].font = Font(bold=True, size=10)

        # Row 7: empty
        # Row 8: Column headers
        headers = ["項次", "品名規格", "單位", "數量", "單價(未稅)", "複價(未稅)", "備註"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Build the hierarchical item structure
        current_row = 9
        section_subtotal_rows = []  # Track (section_name, subtotal_row, first_item_row, last_item_row)
        major_section_subtotal_rows = []  # Track major section subtotal rows for grand total

        if quote_data.sections:
            current_row = self._write_sections(
                ws, quote_data, current_row,
                section_subtotal_rows, major_section_subtotal_rows
            )
        else:
            current_row = self._write_flat_items(
                ws, quote_data, current_row,
                section_subtotal_rows, major_section_subtotal_rows
            )

        # Grand total rows
        current_row += 1  # blank row

        # 未稅合計
        ws.merge_cells(f"A{current_row}:E{current_row}")
        ws[f"A{current_row}"] = "未稅合計"
        ws[f"A{current_row}"].font = Font(bold=True, size=11)
        ws[f"A{current_row}"].alignment = Alignment(horizontal="right")
        if major_section_subtotal_rows:
            formula_parts = "+".join(f"F{r}" for r in major_section_subtotal_rows)
            ws[f"F{current_row}"] = f"={formula_parts}"
        else:
            ws[f"F{current_row}"] = quote_data.summary.subtotal
        ws[f"F{current_row}"].font = Font(bold=True, size=11)
        ws[f"F{current_row}"].number_format = "#,##0"
        for c in range(1, 8):
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        subtotal_row = current_row
        current_row += 1

        # 稅額
        ws.merge_cells(f"A{current_row}:E{current_row}")
        ws[f"A{current_row}"] = "稅額 (5%)"
        ws[f"A{current_row}"].font = Font(bold=True, size=11)
        ws[f"A{current_row}"].alignment = Alignment(horizontal="right")
        ws[f"F{current_row}"] = f"=ROUND(F{subtotal_row}*$E$6,0)"
        ws[f"F{current_row}"].font = Font(bold=True, size=11)
        ws[f"F{current_row}"].number_format = "#,##0"
        for c in range(1, 8):
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        tax_row = current_row
        current_row += 1

        # 含稅合計
        ws.merge_cells(f"A{current_row}:E{current_row}")
        ws[f"A{current_row}"] = "含稅合計"
        ws[f"A{current_row}"].font = Font(bold=True, size=11)
        ws[f"A{current_row}"].alignment = Alignment(horizontal="right")
        ws[f"F{current_row}"] = f"=F{subtotal_row}+F{tax_row}"
        ws[f"F{current_row}"].font = Font(bold=True, size=11)
        ws[f"F{current_row}"].number_format = "#,##0"
        for c in range(1, 8):
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        current_row += 2

        # Signature area
        ws.merge_cells(f"A{current_row}:C{current_row}")
        ws[f"A{current_row}"] = "報價廠商：________________"
        ws.merge_cells(f"D{current_row}:G{current_row}")
        ws[f"D{current_row}"] = "日期：________________"
        current_row += 2
        ws.merge_cells(f"A{current_row}:C{current_row}")
        ws[f"A{current_row}"] = "負責人簽章：________________"
        ws.merge_cells(f"D{current_row}:G{current_row}")
        ws[f"D{current_row}"] = "聯絡電話：________________"

        ws.print_area = f"A1:G{current_row}"

    def _write_sections(self, ws, quote_data, start_row, section_subtotal_rows, major_subtotal_rows):
        """Write items organized by sections."""
        current_row = start_row
        sections = quote_data.sections

        for sec_idx, section in enumerate(sections):
            sec_num = CHINESE_NUMERALS[sec_idx] if sec_idx < len(CHINESE_NUMERALS) else str(sec_idx + 1)

            # Major section header
            ws.merge_cells(f"A{current_row}:G{current_row}")
            ws[f"A{current_row}"] = f"{sec_num} {section.section_name}"
            ws[f"A{current_row}"].font = Font(bold=True, size=11)
            ws[f"A{current_row}"].fill = SECTION_FILL
            ws[f"A{current_row}"].border = THIN_BORDER
            current_row += 1

            # Group items by sub-sections if they exist
            sub_sections = self._detect_sub_sections(section.items)
            first_item_row = current_row

            if sub_sections:
                sub_subtotal_rows = []
                for sub_idx, (sub_name, sub_items) in enumerate(sub_sections):
                    sub_num = SUB_NUMERALS[sub_idx] if sub_idx < len(SUB_NUMERALS) else str(sub_idx + 1)

                    # Sub-section header
                    ws.merge_cells(f"A{current_row}:G{current_row}")
                    ws[f"A{current_row}"] = f"{sub_num} {sub_name}"
                    ws[f"A{current_row}"].font = Font(bold=True, size=10)
                    ws[f"A{current_row}"].border = THIN_BORDER
                    current_row += 1

                    sub_first = current_row
                    for item in sub_items:
                        if item.is_subtotal or item.is_header:
                            continue
                        current_row = self._write_item_row(ws, current_row, item)

                    sub_last = current_row - 1

                    # Sub-section subtotal
                    ws.merge_cells(f"A{current_row}:E{current_row}")
                    ws[f"A{current_row}"] = f"{sub_name}小計"
                    ws[f"A{current_row}"].font = Font(bold=True, size=10)
                    ws[f"A{current_row}"].alignment = Alignment(horizontal="right")
                    ws[f"F{current_row}"] = f"=SUM(F{sub_first}:F{sub_last})"
                    ws[f"F{current_row}"].font = Font(bold=True, size=10)
                    ws[f"F{current_row}"].number_format = "#,##0"
                    for c in range(1, 8):
                        ws.cell(row=current_row, column=c).border = THIN_BORDER
                    sub_subtotal_rows.append(current_row)
                    current_row += 1

                # If there are multiple sub-sections, add a section total
                if len(sub_subtotal_rows) > 1:
                    ws.merge_cells(f"A{current_row}:E{current_row}")
                    ws[f"A{current_row}"] = f"{section.section_name}合計"
                    ws[f"A{current_row}"].font = Font(bold=True, size=11)
                    ws[f"A{current_row}"].alignment = Alignment(horizontal="right")
                    formula_parts = "+".join(f"F{r}" for r in sub_subtotal_rows)
                    ws[f"F{current_row}"] = f"={formula_parts}"
                    ws[f"F{current_row}"].font = Font(bold=True, size=11)
                    ws[f"F{current_row}"].number_format = "#,##0"
                    for c in range(1, 8):
                        ws.cell(row=current_row, column=c).border = THIN_BORDER
                    major_subtotal_rows.append(current_row)
                    current_row += 1
                else:
                    major_subtotal_rows.extend(sub_subtotal_rows)
            else:
                # No sub-sections, just items
                for item in section.items:
                    if item.is_subtotal or item.is_header:
                        continue
                    current_row = self._write_item_row(ws, current_row, item)

                last_item_row = current_row - 1

                # Section subtotal
                ws.merge_cells(f"A{current_row}:E{current_row}")
                ws[f"A{current_row}"] = f"{section.section_name}小計"
                ws[f"A{current_row}"].font = Font(bold=True, size=10)
                ws[f"A{current_row}"].alignment = Alignment(horizontal="right")
                ws[f"F{current_row}"] = f"=SUM(F{first_item_row}:F{last_item_row})"
                ws[f"F{current_row}"].font = Font(bold=True, size=10)
                ws[f"F{current_row}"].number_format = "#,##0"
                for c in range(1, 8):
                    ws.cell(row=current_row, column=c).border = THIN_BORDER
                major_subtotal_rows.append(current_row)
                current_row += 1

        return current_row

    def _write_flat_items(self, ws, quote_data, start_row, section_subtotal_rows, major_subtotal_rows):
        """Write items without section structure."""
        current_row = start_row
        first_item_row = current_row

        for item in quote_data.all_items:
            if item.is_header:
                ws.merge_cells(f"A{current_row}:G{current_row}")
                ws[f"A{current_row}"] = item.description
                ws[f"A{current_row}"].font = Font(bold=True, size=10)
                ws[f"A{current_row}"].fill = SECTION_FILL
                ws[f"A{current_row}"].border = THIN_BORDER
                current_row += 1
            elif item.is_subtotal:
                last_item_row = current_row - 1
                ws.merge_cells(f"A{current_row}:E{current_row}")
                ws[f"A{current_row}"] = item.description
                ws[f"A{current_row}"].font = Font(bold=True, size=10)
                ws[f"A{current_row}"].alignment = Alignment(horizontal="right")
                ws[f"F{current_row}"] = f"=SUM(F{first_item_row}:F{last_item_row})"
                ws[f"F{current_row}"].font = Font(bold=True, size=10)
                ws[f"F{current_row}"].number_format = "#,##0"
                for c in range(1, 8):
                    ws.cell(row=current_row, column=c).border = THIN_BORDER
                major_subtotal_rows.append(current_row)
                first_item_row = current_row + 1
                current_row += 1
            else:
                current_row = self._write_item_row(ws, current_row, item)

        return current_row

    def _write_item_row(self, ws, row, item: QuoteLineItem) -> int:
        """Write a single item row with formula. Returns next row number."""
        desc = item.description
        if item.specification:
            desc += f"\n{item.specification}"

        row_data = [
            (1, item.item_code or str(item.seq)),
            (2, desc),
            (3, item.unit),
            (4, item.quantity if item.quantity else ""),
        ]

        for col, val in row_data:
            cell = ws.cell(row=row, column=col)
            cell.value = val
            cell.border = THIN_BORDER
            cell.font = Font(size=10)
            if col == 2:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col == 4:
                cell.alignment = Alignment(horizontal="right")

        # E column: blank (vendor fills in)
        ws.cell(row=row, column=5).border = THIN_BORDER
        ws.cell(row=row, column=5).number_format = "#,##0"

        # F column: formula
        ws.cell(row=row, column=6).value = f'=IF(OR(D{row}="",E{row}=""),"",D{row}*E{row})'
        ws.cell(row=row, column=6).border = THIN_BORDER
        ws.cell(row=row, column=6).number_format = "#,##0"

        # G column: remark (blank)
        ws.cell(row=row, column=7).border = THIN_BORDER

        return row + 1

    def _detect_sub_sections(self, items: list[QuoteLineItem]) -> list[tuple[str, list[QuoteLineItem]]]:
        """Detect sub-section groupings within items based on item codes."""
        # Check if items have grouping patterns (e.g., 1.x, 2.x, etc.)
        # For now, return empty to treat all items as flat within section
        return []

    def _build_verification_sheet(self, ws, quote_data: QuoteData):
        """Build the _驗證 verification sheet."""
        meta = quote_data.metadata

        # Column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 45
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 14
        ws.column_dimensions["J"].width = 8
        ws.column_dimensions["L"].width = 20
        ws.column_dimensions["M"].width = 25

        # Row 1-4: Header
        ws["A1"] = "來源文件"
        ws["B1"] = f"{meta.document_number or ''}-{meta.project_name}"
        ws["A2"] = "工程名稱"
        ws["B2"] = meta.project_name
        ws["A3"] = "報價單號"
        ws["B3"] = meta.document_number or ""
        ws["A4"] = "報價日期"
        ws["B4"] = str(meta.quote_date).replace("-", "/") if meta.quote_date else ""

        for r in range(1, 5):
            ws[f"A{r}"].font = Font(bold=True, size=10)

        # Verification summary (right side)
        ws["L1"] = "驗證摘要"
        ws["L1"].font = Font(bold=True, size=10)

        # Row 6: Table headers
        v_headers = ["節", "分項", "項次", "品名規格", "單位", "數量",
                      "原單價", "原金額", "數量×原單價", "檢核"]
        for col_idx, header in enumerate(v_headers, 1):
            cell = ws.cell(row=6, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        # Data rows
        current_row = 7
        mismatch_count = 0
        items = quote_data.get_real_items()

        for item in items:
            section_name = item.section or ""
            calc_price = item.quantity * item.unit_price if item.unit_price else 0
            check_result = "OK"
            if item.total_price and item.unit_price and item.quantity:
                if abs(calc_price - item.total_price) > 1:
                    check_result = f"不一致：{item.quantity} × {item.unit_price:,.0f} ≠ {item.total_price:,.0f}"
                    mismatch_count += 1

            row_data = [
                section_name,
                "",  # sub-section
                item.item_code or str(item.seq),
                item.description,
                item.unit,
                item.quantity,
                item.unit_price,
                item.total_price,
                calc_price,
                check_result,
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                cell.border = THIN_BORDER
                cell.font = Font(size=9)
                if col_idx in (7, 8, 9):
                    cell.number_format = "#,##0"

            if check_result != "OK":
                for col_idx in range(1, 11):
                    ws.cell(row=current_row, column=col_idx).font = Font(
                        size=9, color="FF0000"
                    )

            current_row += 1

        # Fill verification summary
        subtotal = sum(it.total_price for it in items)
        ws["L2"] = "細項算術不一致筆數"
        ws["M2"] = mismatch_count
        ws["L3"] = "項目金額合計"
        ws["M3"] = subtotal
        ws["M3"].number_format = "#,##0"
        ws["L4"] = "原文件合計"
        ws["M4"] = quote_data.summary.subtotal
        ws["M4"].number_format = "#,##0"
        ws["L6"] = "總計(未稅)"
        ws["M6"] = quote_data.summary.subtotal
        ws["M6"].number_format = "#,##0"

        if mismatch_count:
            ws["L7"] = "驗證結果"
            ws["M7"] = f"來源文件存在 {mismatch_count} 筆細項算術不一致"
            ws["M7"].font = Font(size=9, color="FF0000")
        else:
            ws["L7"] = "驗證結果"
            ws["M7"] = "所有細項算術一致"
