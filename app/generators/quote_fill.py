"""Quote Fill Generator (Excel output).

Generates the 欣興 internal requisition form matching the sample format:
- Sheet 1 "(空白)保養、維修、改善、工程說明": Form with merged cells, AI narratives
- Sheet 2 "請勿刪除": Dropdown validation data
"""

import io
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from app.generators.base import BaseGenerator
from app.models.quote_data import QuoteData, QuoteFillContent

logger = logging.getLogger(__name__)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
LABEL_FONT = Font(bold=True, size=10)
VALUE_FONT = Font(size=10)
LABEL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


class QuoteFillGenerator(BaseGenerator):
    doc_type = "quote_fill"
    output_extension = "xlsx"

    async def generate(self, quote_data: QuoteData, **kwargs) -> bytes:
        """Generate filled 欣興 internal form."""
        fill_content = kwargs.get("quote_fill_content")
        if not fill_content and self.llm:
            fill_content = await self.llm.generate_quote_fill_content(quote_data)
        if not fill_content:
            from app.llm.batch_processor import _fallback_quote_fill_content
            fill_content = _fallback_quote_fill_content(quote_data)

        if self.template_path and self.template_path.exists():
            return self._fill_template(quote_data, fill_content)
        else:
            return self._generate_from_scratch(quote_data, fill_content)

    def _fill_template(self, quote_data: QuoteData, content: QuoteFillContent) -> bytes:
        """Fill an existing Excel template."""
        from app.templates.engine import ExcelTemplateEngine
        engine = ExcelTemplateEngine(self.template_path)
        return engine.save_to_bytes()

    def _generate_from_scratch(self, quote_data: QuoteData, content: QuoteFillContent) -> bytes:
        """Generate the complete 欣興 form from scratch."""
        wb = openpyxl.Workbook()

        # Sheet 1: Form
        ws = wb.active
        ws.title = "(空白)保養、維修、改善、工程說明"
        self._build_form_sheet(ws, quote_data, content)

        # Sheet 2: Dropdown data
        ws2 = wb.create_sheet("請勿刪除")
        self._build_dropdown_sheet(ws2)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _build_form_sheet(self, ws, quote_data: QuoteData, content: QuoteFillContent):
        """Build the main form sheet matching the sample layout."""
        # Page setup
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1

        # Column widths
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].width = 16

        grand_total = quote_data.summary.grand_total

        # ===== Row 1: Header =====
        ws.merge_cells("A1:E1")
        ws["A1"] = "環廠處梅竹區 * * 廠 * * 課"
        ws["A1"].font = Font(bold=True, size=11)

        ws.merge_cells("F1:G1")
        ws["F1"] = "必填(下拉式選單)"
        ws["F1"].font = Font(size=9, color="FF0000")

        ws.merge_cells("H1:I1")
        ws["H1"] = "說明"
        ws["H1"].font = Font(size=9)

        # ===== Row 2: 請購名稱 =====
        ws["A2"] = "*請購名稱"
        ws["A2"].font = LABEL_FONT
        ws["A2"].fill = LABEL_FILL
        ws["A2"].border = THIN_BORDER

        ws.merge_cells("B2:I2")
        ws["B2"] = content.purchase_name
        ws["B2"].font = VALUE_FONT
        ws["B2"].border = THIN_BORDER
        ws["B2"].alignment = Alignment(wrap_text=True)

        # ===== Row 3: 設備名稱, 站別, 位置, 施作總成本 =====
        _label(ws, "A3", "*設備名稱")
        _value(ws, "B3", content.equipment_name)
        _label(ws, "C3", "*站別")
        _value(ws, "D3", content.station)
        _label(ws, "E3", "*位置")
        _value(ws, "F3", content.location)
        ws.merge_cells("G3:H3")
        _label(ws, "G3", "*施作總成本(請購金額)")
        _value(ws, "I3", f"{grand_total:,.0f}")

        # ===== Row 4: 費用別, 請購類別 =====
        _label(ws, "A4", "*費用別")
        _value(ws, "B4", content.cost_type)
        ws.merge_cells("C4:D4")
        _label(ws, "C4", "*請購類別(修繕)")
        ws.merge_cells("E4:F4")
        _value(ws, "E4", content.purchase_category_repair)
        ws.merge_cells("G4:H4")
        _label(ws, "G4", "*請購類別(非修繕)")
        _value(ws, "I4", content.purchase_category_non_repair)

        # ===== Row 5: 施作原因, 使用週期, 合理週期, 前次施作 =====
        _label(ws, "A5", "*施作原因")
        _value(ws, "B5", content.work_reason)
        _label(ws, "C5", "*使用週期(月)")
        _value(ws, "D5", "NA(新設/改善工程)")
        _label(ws, "E5", "*合理週期(月)")
        _value(ws, "F5", "NA(新設/改善工程)")
        ws.merge_cells("G5:H5")
        _label(ws, "G5", "*前次施作(西元年/月)")
        _value(ws, "I5", "NA(本次新設/改善案)")

        # ===== Row 6-7: 現況說明:1 =====
        ws.merge_cells("A6:A7")
        _label(ws, "A6", "*現況說明:1\n(現況把握)")
        ws["A6"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("B6:C7")
        _value(ws, "B6", content.situation_desc_1)
        ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("D6:D7")
        _label(ws, "D6", "*照片/施作位置：")
        ws["D6"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("E6:F7")
        _value(ws, "E6", content.photo_location_1a)
        ws["E6"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("G6:G7")
        _label(ws, "G6", "*照片/施作位置：")
        ws["G6"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("H6:I7")
        _value(ws, "H6", content.photo_location_1b)
        ws["H6"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[6].height = 50
        ws.row_dimensions[7].height = 30

        # ===== Row 8: 現況問題及風險 =====
        _label(ws, "A8", "*現況問題及風險\n(原因分析)")
        ws["A8"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("B8:C8")
        _value(ws, "B8", content.problem_risk_1)
        ws["B8"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("D8:F10")
        ws["D8"] = "(張貼照片)"
        ws["D8"].font = Font(size=9, color="808080")
        ws["D8"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("G8:I10")
        ws["G8"] = "(張貼照片)"
        ws["G8"].font = Font(size=9, color="808080")
        ws["G8"].alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[8].height = 55

        # ===== Row 9: 對策評估說明 =====
        _label(ws, "A9", "*對策評估說明\n(真因證實/對策擬定)")
        ws["A9"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("B9:C9")
        _value(ws, "B9", content.countermeasure_1)
        ws["B9"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[9].height = 55

        # ===== Row 10: 執行改善/維修對策 =====
        _label(ws, "A10", "*執行改善/維修對策:\n(對策實施)")
        ws["A10"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("B10:C10")
        _value(ws, "B10", content.execution_1)
        ws["B10"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[10].height = 55

        # ===== Row 11-12: 現況說明:2 =====
        ws.merge_cells("A11:A12")
        _label(ws, "A11", "*現況說明:2\n(現況把握)")
        ws["A11"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("B11:C12")
        _value(ws, "B11", content.situation_desc_2)
        ws["B11"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("D11:D12")
        _label(ws, "D11", "*照片/施作位置：")
        ws["D11"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("E11:F12")
        _value(ws, "E11", content.photo_location_2a)
        ws["E11"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("G11:G12")
        _label(ws, "G11", "*照片/施作位置：")
        ws["G11"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("H11:I12")
        _value(ws, "H11", content.photo_location_2b)
        ws["H11"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[11].height = 50
        ws.row_dimensions[12].height = 30

        # ===== Row 13: 現況問題及風險 2 =====
        _label(ws, "A13", "*現況問題及風險\n(原因分析)")
        ws["A13"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("B13:C13")
        _value(ws, "B13", content.problem_risk_2)
        ws["B13"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("D13:F15")
        ws["D13"] = "(張貼照片)"
        ws["D13"].font = Font(size=9, color="808080")
        ws["D13"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("G13:I15")
        ws["G13"] = "(張貼照片)"
        ws["G13"].font = Font(size=9, color="808080")
        ws["G13"].alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[13].height = 55

        # ===== Row 14: 改善/維修說明 =====
        _label(ws, "A14", "*改善/維修說明:\n(對策實施)")
        ws["A14"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("B14:C14")
        _value(ws, "B14", content.improvement_desc)
        ws["B14"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[14].height = 55

        # ===== Row 15: 執行改善/維修對策 2 =====
        _label(ws, "A15", "*執行改善/維修對策:\n(對策實施)")
        ws["A15"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.merge_cells("B15:C15")
        _value(ws, "B15", content.execution_2)
        ws["B15"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[15].height = 55

        # ===== Row 16: 補充說明 =====
        _label(ws, "A16", "補充說明")

        ws.merge_cells("B16:I16")
        _value(ws, "B16", content.supplementary_notes)
        ws["B16"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[16].height = 60

        # ===== Row 17: Version =====
        from datetime import date as d
        ws["I17"] = f"{d.today().strftime('%Y.%m.%d')}版"
        ws["I17"].font = Font(size=8, color="808080")

        # Apply borders to all form cells
        for row in range(1, 17):
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                if cell.border == Border():
                    cell.border = THIN_BORDER

    def _build_dropdown_sheet(self, ws):
        """Build the dropdown validation data sheet."""
        dropdown_data = {
            "A": ["費用別", "必填(下拉式選單)", "修繕", "非修繕"],
            "B": ["請購類別(修繕)", "NA(下拉式選單)", "PM(定保)", "BM(維修)",
                   "CM(改善)", "工程", "房屋建築", "合約", "環保", "其他"],
            "C": ["請購類別(非修繕)", "NA(下拉式選單)", "新設工程(20萬以上)",
                   "移機運費", "物料", "校驗費", "清潔費", "雜項/其他"],
            "D": ["施作原因", "必填(下拉式選單)", "定期保養", "點檢檢出",
                   "正常損壞", "人為損壞", "品質異常", "設計缺失",
                   "降低成本", "移機裝機", "一二次配", "保養合約",
                   "節水節電", "安全疑慮", "其他"],
            "E": ["週期", "必填(下拉式選單)", "1個月", "2個月", "3個月",
                   "4個月", "5個月", "6個月", "7個月", "8個月",
                   "9個月", "10個月", "11個月", "12個月"],
            "F": ["廠別", "必填(下拉式選單)", "H02", "H05", "O02",
                   "HQ", "P", "CUB", "其他"],
            "G": ["課別", "必填(下拉式選單)", "本部", "公設一課",
                   "公設二課", "公設三課", "環工一課", "環工二課", "環工三課"],
            "H": ["類別", "必填(下拉式選單)", "設備保養", "設備維修",
                   "設備改善", "工程", "其他"],
            "I": ["待料時間", "NA(下拉式選單)"] + [str(i) for i in range(1, 13)],
            "J": ["改善幅度", "必填(下拉式選單)", "提升", "降低"],
        }

        for col_letter, values in dropdown_data.items():
            for row_idx, val in enumerate(values, 1):
                ws[f"{col_letter}{row_idx}"] = val
                if row_idx == 1:
                    ws[f"{col_letter}{row_idx}"].font = Font(bold=True, size=10)


def _label(ws, cell_ref: str, text: str):
    """Write a label cell."""
    ws[cell_ref] = text
    ws[cell_ref].font = LABEL_FONT
    ws[cell_ref].fill = LABEL_FILL
    ws[cell_ref].border = THIN_BORDER


def _value(ws, cell_ref: str, text):
    """Write a value cell."""
    ws[cell_ref] = text
    ws[cell_ref].font = VALUE_FONT
    ws[cell_ref].border = THIN_BORDER
